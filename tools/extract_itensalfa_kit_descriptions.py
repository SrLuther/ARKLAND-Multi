#!/usr/bin/env python3
"""Extrai características dos itens ItensAlfa da planilha para a UI de kits.

Lê materiais de craft (ASE Armaduras / Armas / Ferramentas) e stats por tier
(Status dos itens), indexa por blueprint e grava JSON consumido pelo
catalog_enrich da Web Store.

Uso:
  python tools/extract_itensalfa_kit_descriptions.py
  python tools/extract_itensalfa_kit_descriptions.py --xlsx "C:/Users/Ciano/Downloads/Itens Alfa.xlsx"
"""
from __future__ import annotations

import argparse
import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path(r"C:\Users\Ciano\Downloads\Itens Alfa.xlsx")
OUT_TOOLS = ROOT / "tools" / "itensalfa_kit_descriptions.json"
OUT_WEB = ROOT / "plugin" / "arkshop_web" / "data" / "itensalfa_kit_descriptions.json"

TIER_ALIASES = {
    "delta": "Delta",
    "gama": "Gama",
    "gamma": "Gama",
    "beta": "Beta",
    "alfa": "Alfa",
    "alpha": "Alfa",
    "omega": "Omega",
    "transcendente": "Transcendente",
    "etereo": "Etereo",
    "etéreo": "Etereo",
    "universal": "Universal",
    "onipotente": "Onipotente",
    "surreal": "Surreal",
    "imaterial": "Imaterial",
    "exotico": "Exotico",
    "exótico": "Exotico",
}

KIND_STATUS_KEY = {
    "armor": "armor",
    "weapon": "weapon",
    "tool": None,  # planilha Status não tem coluna de ferramentas
    "saddle": "saddle",
}

FRIENDLY_OVERRIDES: dict[str, str] = {
    "TEK SHIELD ARMOR": "Escudo TEK",
    "SHOULDER CANNON / CANHÃO DE OMBRO": "Canhão de Ombro",
    "SHOULDER CANNON / CANHAO DE OMBRO": "Canhão de Ombro",
    "TEK BOW": "Arco TEK",
    "TEK PISTOL": "Pistola TEK",
    "ELECTROPOD": "ElectroPod",
    "TEK SWORD": "Espada TEK",
    "TEK CLAWS / GARRAS TEK": "Garras TEK",
    "TEK RIFLE": "Rifle TEK",
    "SNIPER": "Sniper",
    "PIKE": "Lança",
    "PUMP-ACTION": "Espingarda Pump-Action",
    "CLUB  /CLAVA": "Clava",
    "CLUB /CLAVA": "Clava",
    "TEK GRENADE LAUNCHER": "Lançador de Granadas TEK",
    "TEK CRUISE MISSIL": "Míssil de Cruzeiro TEK",
    "CHAINSAW / MOTOSSERRA": "Motosserra",
    "HATCHED / MACHADO": "Machado",
    "MININGDRILL": "Perfuratriz",
    "PICK / PICARETA": "Picareta",
    "SICKLE / FOICE": "Foice",
    "FISHING ROD / VARA DE PESCA": "Vara de Pesca",
    "TORCH / TOCHA": "Tocha",
    "WHIP / CHICOTE": "Chicote",
    "LANTERN CHARGE / LANTERNA DE CARGA": "Lanterna de Carga",
    "BOTAS TEK": "Botas TEK",
    "LUVAS TEK": "Luvas TEK",
    "CAPACETE TEK": "Capacete TEK",
    "CALÇAS TEK": "Calças TEK",
    "CALCAS TEK": "Calças TEK",
    "PEITORAL TEK": "Peitoral TEK",
}


def _norm(text: str) -> str:
    return (
        unicodedata.normalize("NFD", text or "")
        .encode("ascii", "ignore")
        .decode()
        .lower()
        .strip()
    )


def _norm_tier(raw: Any) -> str | None:
    if raw is None:
        return None
    key = _norm(str(raw))
    return TIER_ALIASES.get(key)


def _extract_blueprint(cell: Any) -> str | None:
    if not isinstance(cell, str) or "Blueprint" not in cell:
        return None
    m = re.search(r"Blueprint'([^']+)'", cell)
    if not m:
        m = re.search(r'Blueprint"([^"]+)"', cell)
    if not m:
        return None
    bp = m.group(1).strip()
    if "." not in bp.rsplit("/", 1)[-1]:
        # GiveItem sometimes omits class suffix — keep as-is
        pass
    return bp


def _friendly_name(title: str) -> str:
    cleaned = re.sub(r"\s+", " ", (title or "").strip())
    if cleaned in FRIENDLY_OVERRIDES:
        return FRIENDLY_OVERRIDES[cleaned]
    key = cleaned.upper()
    for k, v in FRIENDLY_OVERRIDES.items():
        if _norm(k) == _norm(cleaned):
            return v
    # fallback: take PT side after /
    if "/" in cleaned:
        parts = [p.strip() for p in cleaned.split("/") if p.strip()]
        if len(parts) >= 2:
            return parts[-1].title() if parts[-1].isupper() else parts[-1]
    return cleaned.title() if cleaned.isupper() else cleaned


def _is_tier_row(first: Any) -> bool:
    return _norm_tier(first) is not None


def _is_material_header(row: list[Any]) -> bool:
    """Header de materiais: 1ª célula vazia + nomes de recurso nas seguintes."""
    if len(row) < 2:
        return False
    first = row[0]
    if first is not None and str(first).strip() != "":
        return False
    labels = [
        str(c).strip()
        for c in row[1:]
        if c is not None and str(c).strip() and not str(c).startswith("---")
    ]
    if len(labels) < 2:
        return False
    # Evitar confundir com linhas de cheat/separador
    if any("Blueprint" in lab or "GiveItem" in lab for lab in labels):
        return False
    return True


def _parse_amount(val: Any) -> int | None:
    if val is None or val == "" or val == "-":
        return None
    try:
        return int(round(float(val)))
    except (TypeError, ValueError):
        return None


def _parse_sheet_blocks(
    rows: list[list[Any]],
    *,
    kind: str,
) -> list[dict[str, Any]]:
    """Parse blocks: TITLE, material headers, tier rows with optional cheat BP."""
    entries: list[dict[str, Any]] = []
    i = 0
    n = len(rows)
    while i < n:
        row = rows[i]
        first = row[0] if row else None
        if not (isinstance(first, str) and first.strip()) or _is_tier_row(first):
            i += 1
            continue
        if _is_material_header(row):
            i += 1
            continue

        title = first.strip()
        # skip section banners like "TODAS AS ARMADURAS TEK"
        if title.upper().startswith("TODAS AS"):
            i += 1
            continue

        headers: list[str] = []
        j = i + 1
        if j < n and _is_material_header(rows[j]):
            headers = [
                str(c).strip()
                for c in rows[j][1:]
                if c is not None and str(c).strip() and not str(c).startswith("---")
            ]
            # drop trailing separator-looking headers
            headers = [h for h in headers if not set(h) <= {"-", "–"}]
            j += 1

        while j < n:
            r = rows[j]
            tier = _norm_tier(r[0] if r else None)
            if tier is None:
                # next title or blank
                if isinstance(r[0], str) and r[0].strip() and not _is_material_header(r):
                    break
                if r[0] is None and all(c is None or c == "" for c in r):
                    j += 1
                    continue
                break

            materials: list[dict[str, Any]] = []
            for idx, h in enumerate(headers):
                col = idx + 1
                amt = _parse_amount(r[col] if col < len(r) else None)
                if amt is None or amt <= 0:
                    continue
                materials.append({"name": h, "amount": amt})

            bp = None
            for cell in r:
                bp = _extract_blueprint(cell)
                if bp:
                    break

            if bp:
                entries.append({
                    "blueprint": bp,
                    "name": _friendly_name(title),
                    "title_raw": title,
                    "kind": kind,
                    "tier": tier,
                    "materials": materials,
                })
            j += 1
        i = j
    return entries


def _load_status(ws_rows: list[list[Any]]) -> dict[str, dict[str, int]]:
    out: dict[str, dict[str, int]] = {}
    for row in ws_rows[1:]:
        tier = _norm_tier(row[0] if row else None)
        if not tier:
            continue
        armor = _parse_amount(row[1] if len(row) > 1 else None)
        weapon = _parse_amount(row[2] if len(row) > 2 else None)
        saddle = _parse_amount(row[3] if len(row) > 3 else None)
        entry: dict[str, int] = {}
        if armor is not None:
            entry["armor"] = armor
        if weapon is not None:
            entry["weapon"] = weapon
        if saddle is not None:
            entry["saddle"] = saddle
        if entry:
            out[tier] = entry
    return out


def _stats_for(kind: str, tier: str, status: dict[str, dict[str, int]]) -> dict[str, Any]:
    tier_stats = status.get(tier) or {}
    key = KIND_STATUS_KEY.get(kind)
    if not key or key not in tier_stats:
        return {}
    val = tier_stats[key]
    if kind == "armor":
        return {"armor": val, "label": f"Armadura {val}"}
    if kind == "weapon":
        return {"damage": val, "label": f"Dano {val}"}
    if kind == "saddle":
        return {"armor": val, "label": f"Armadura da sela {val}"}
    return {}


def _materials_text(materials: list[dict[str, Any]]) -> str:
    if not materials:
        return ""
    return ", ".join(f"{m['name']}×{m['amount']}" for m in materials)


def _summary(
    name: str,
    kind: str,
    tier: str,
    stats: dict[str, Any],
    materials: list[dict[str, Any]],
) -> str:
    kind_pt = {
        "armor": "Armadura",
        "weapon": "Arma",
        "tool": "Ferramenta",
        "saddle": "Sela",
    }.get(kind, kind)
    parts = [f"{kind_pt} {tier}"]
    if stats.get("label"):
        parts.append(str(stats["label"]))
    mats = _materials_text(materials)
    if mats:
        parts.append(f"Craft: {mats}")
    return " · ".join(parts)


def _sheet_rows(wb: Any, name: str) -> list[list[Any]]:
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def extract(xlsx: Path) -> dict[str, Any]:
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit(
            "openpyxl é necessário: pip install openpyxl"
        ) from exc

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    status = _load_status(_sheet_rows(wb, "Status dos itens"))

    blocks: list[dict[str, Any]] = []
    blocks.extend(_parse_sheet_blocks(_sheet_rows(wb, "ASE - Armaduras"), kind="armor"))
    blocks.extend(_parse_sheet_blocks(_sheet_rows(wb, "ASE Armas"), kind="weapon"))
    blocks.extend(_parse_sheet_blocks(_sheet_rows(wb, "ASE - Ferramentas"), kind="tool"))
    # Selas sheet may exist with similar layout
    if "Selas" in wb.sheetnames:
        blocks.extend(_parse_sheet_blocks(_sheet_rows(wb, "Selas"), kind="saddle"))

    by_bp: dict[str, dict[str, Any]] = {}
    for block in blocks:
        bp = block["blueprint"]
        stats = _stats_for(block["kind"], block["tier"], status)
        mats = block.get("materials") or []
        entry = {
            "name": block["name"],
            "kind": block["kind"],
            "tier": block["tier"],
            "stats": stats,
            "materials": mats,
            "materials_text": _materials_text(mats),
            "summary": _summary(block["name"], block["kind"], block["tier"], stats, mats),
        }
        # Prefer entries with materials if duplicate BP
        prev = by_bp.get(bp)
        if prev is None or (mats and not prev.get("materials")):
            by_bp[bp] = entry

    return {
        "source": str(xlsx),
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "notes": (
            "Características ItensAlfa: stats (Status dos itens) + materiais de craft "
            "(ASE Armaduras/Armas/Ferramentas). Indexado por blueprint para kits kit_itensalfa_*."
        ),
        "status_by_tier": status,
        "by_blueprint": by_bp,
        "count": len(by_bp),
    }


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    if not args.xlsx.is_file():
        raise SystemExit(f"Planilha não encontrada: {args.xlsx}")

    data = extract(args.xlsx)
    print(f"Extraídos {data['count']} blueprints de {args.xlsx.name}")
    # sample
    sample_keys = list(data["by_blueprint"])[:3]
    for k in sample_keys:
        name = data["by_blueprint"][k]["name"]
        summary = data["by_blueprint"][k]["summary"][:100]
        print(f"  {name}: {summary}")

    text = json.dumps(data, ensure_ascii=False, indent=2) + "\n"
    if args.dry_run:
        print("(dry-run — não gravou)")
        return
    OUT_TOOLS.parent.mkdir(parents=True, exist_ok=True)
    OUT_WEB.parent.mkdir(parents=True, exist_ok=True)
    OUT_TOOLS.write_text(text, encoding="utf-8")
    OUT_WEB.write_text(text, encoding="utf-8")
    print(f"Gravado: {OUT_TOOLS.relative_to(ROOT)}")
    print(f"Gravado: {OUT_WEB.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
